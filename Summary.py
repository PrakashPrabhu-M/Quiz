# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Summary.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook
from PyQt5.QtWidgets import QMessageBox

class Ui_Result(object):
    def view(self,name):
        try:
            self.textBrowser.clear()        
            active=self.wb[name]
            data=[i for i in active.values]
            self.textBrowser.append(data[0][0]+'\t'+data[0][1])
            for i in data[1:]:
                self.textBrowser.append(str(i[0])+'\t\t'+str(i[1]))
              
            total=0
            n=0
            for i in data[1:]:
                n+=1
                total+=int(i[1])
            percentage=total/n*10
            self.textBrowser.append('\n\nPercentage => {}% '.format(percentage))
        except:
            msg=QMessageBox()
            msg.setWindowTitle('Info')
            msg.setText('Error!')
            msg.setIcon(QMessageBox.Critical)
            x=msg.exec_()
            
    def setupUi(self, Result):
        self.wb=load_workbook(r'Users\Accounts.xlsx')
        Result.setObjectName("Result")
        Result.resize(810, 682)
        self.centralwidget = QtWidgets.QWidget(Result)
        self.centralwidget.setObjectName("centralwidget")
        self.L_id = QtWidgets.QLabel(self.centralwidget)
        self.L_id.setGeometry(QtCore.QRect(40, 30, 181, 81))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(22)
        font.setBold(True)
        font.setWeight(75)
        self.L_id.setFont(font)
        self.L_id.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.L_id.setObjectName("L_id")
        self.L_result = QtWidgets.QLabel(self.centralwidget)
        self.L_result.setGeometry(QtCore.QRect(50, 120, 181, 81))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(22)
        font.setBold(True)
        font.setWeight(75)
        self.L_result.setFont(font)
        self.L_result.setObjectName("L_result")
        self.textBrowser = QtWidgets.QTextBrowser(self.centralwidget)
        self.textBrowser.setGeometry(QtCore.QRect(220, 140, 521, 441))
        self.textBrowser.viewport().setProperty("cursor", QtGui.QCursor(QtCore.Qt.BlankCursor))
        self.textBrowser.setObjectName("textBrowser")
        self.B_view = QtWidgets.QPushButton(self.centralwidget)
        self.B_view.setGeometry(QtCore.QRect(630, 50, 111, 41))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.B_view.setFont(font)
        self.B_view.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.B_view.setObjectName("B_view")
        
        self.B_view.clicked.connect(lambda:self.view(self.comboBox.currentText()))
        
        self.B_close = QtWidgets.QPushButton(self.centralwidget)
        self.B_close.setGeometry(QtCore.QRect(680, 590, 93, 28))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        self.B_close.setFont(font)
        self.B_close.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.B_close.setObjectName("B_close")
        
        self.B_close.clicked.connect(Result.close)
        
        self.bg_img = QtWidgets.QLabel(self.centralwidget)
        self.bg_img.setGeometry(QtCore.QRect(4, 0, 801, 651))
        self.bg_img.setText("")
        self.bg_img.setPixmap(QtGui.QPixmap(r"Assets/Result/background.png"))
        self.bg_img.setScaledContents(True)
        self.bg_img.setObjectName("bg_img")
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(220, 50, 381, 51))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(22)
        self.comboBox.setFont(font)
        self.comboBox.setObjectName("comboBox")
        
        for i in range(len(self.wb.sheetnames)):
            self.comboBox.addItem("")
            self.comboBox.addItem("")
            self.comboBox.addItem("")
            
        self.bg_img.raise_()
        self.L_id.raise_()
        self.L_result.raise_()
        self.textBrowser.raise_()
        self.B_view.raise_()
        self.B_close.raise_()
        self.comboBox.raise_()
        Result.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(Result)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 810, 26))
        self.menubar.setObjectName("menubar")
        Result.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(Result)
        self.statusbar.setObjectName("statusbar")
        Result.setStatusBar(self.statusbar)

        self.retranslateUi(Result)
        QtCore.QMetaObject.connectSlotsByName(Result)

    def retranslateUi(self, Result):
        _translate = QtCore.QCoreApplication.translate
        Result.setWindowTitle(_translate("Result", "Summary"))
        self.L_id.setText(_translate("Result", "Accounts:"))
        self.L_result.setText(_translate("Result", "Result:"))
        self.B_view.setToolTip(_translate("Result", "click to view"))
        self.B_view.setText(_translate("Result", "View"))
        self.B_close.setToolTip(_translate("Result", "exit"))
        self.B_close.setText(_translate("Result", "Close"))
        for n,i in enumerate(self.wb.sheetnames[1:]):
            self.comboBox.setItemText(n, _translate("Result", i))
            #self.comboBox.setItemText(1, _translate("Result", i))
            #self.comboBox.setItemText(2, _translate("Result", i))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Result = QtWidgets.QMainWindow()
    ui = Ui_Result()
    ui.setupUi(Result)
    Result.show()
    sys.exit(app.exec_())

