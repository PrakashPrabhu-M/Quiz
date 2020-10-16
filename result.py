# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'result.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook
from PyQt5.QtWidgets import QMessageBox

class Ui_Result(object):
    def view(self,num):
        try:
            self.textBrowser.clear()
            num=int(num)
            wb=load_workbook(r'Users\Accounts.xlsx')
            active=wb[wb.sheetnames[num]]
            data=[i for i in active.values]
            self.textBrowser.append(data[0][0]+'\t'+data[0][1])
            for i in data[1:]:
                self.textBrowser.append(str(i[0])+'\t\t'+str(i[1]))
        except:
            msg=QMessageBox()
            msg.setWindowTitle('Info')
            msg.setText('Error!')
            msg.setIcon(QMessageBox.Critical)
            x=msg.exec_()
        
    
    def setupUi(self, Result):
        Result.setObjectName("Result")
        Result.resize(758, 596)
        self.centralwidget = QtWidgets.QWidget(Result)
        self.centralwidget.setObjectName("centralwidget")
        self.L_id = QtWidgets.QLabel(self.centralwidget)
        self.L_id.setGeometry(QtCore.QRect(40, 30, 181, 81))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(22)
        font.setBold(True)
        font.setWeight(75)
        self.L_id.setFont(font)
        self.L_id.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.L_id.setObjectName("L_id")
        self.LE_id = QtWidgets.QLineEdit(self.centralwidget)
        self.LE_id.setGeometry(QtCore.QRect(220, 40, 351, 61))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(22)
        font.setBold(True)
        font.setWeight(75)
        self.LE_id.setFont(font)
        self.LE_id.setCursor(QtGui.QCursor(QtCore.Qt.IBeamCursor))
        self.LE_id.setObjectName("LE_id")
        self.L_result = QtWidgets.QLabel(self.centralwidget)
        self.L_result.setGeometry(QtCore.QRect(50, 120, 181, 81))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(22)
        font.setBold(True)
        font.setWeight(75)
        self.L_result.setFont(font)
        self.L_result.setObjectName("L_result")
        self.textBrowser = QtWidgets.QTextBrowser(self.centralwidget)
        self.textBrowser.setGeometry(QtCore.QRect(220, 150, 461, 351))
        self.textBrowser.viewport().setProperty("cursor", QtGui.QCursor(QtCore.Qt.BlankCursor))
        self.textBrowser.setObjectName("textBrowser")
        self.B_view = QtWidgets.QPushButton(self.centralwidget)
        self.B_view.setGeometry(QtCore.QRect(590, 47, 111, 41))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.B_view.setFont(font)
        self.B_view.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.B_view.setObjectName("B_view")
        
        self.B_view.clicked.connect(lambda: self.view(self.LE_id.text()))
        
        self.B_close = QtWidgets.QPushButton(self.centralwidget)
        self.B_close.setGeometry(QtCore.QRect(630, 510, 93, 28))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        self.B_close.setFont(font)
        self.B_close.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.B_close.setObjectName("B_close")
        
        self.B_close.clicked.connect(Result.close)
        
        self.bg_img = QtWidgets.QLabel(self.centralwidget)
        self.bg_img.setGeometry(QtCore.QRect(4, 0, 751, 551))
        self.bg_img.setText("")
        self.bg_img.setPixmap(QtGui.QPixmap(r"Assets\Result\background.png"))
        self.bg_img.setScaledContents(True)
        self.bg_img.setObjectName("bg_img")
        self.bg_img.raise_()
        self.L_id.raise_()
        self.LE_id.raise_()
        self.L_result.raise_()
        self.textBrowser.raise_()
        self.B_view.raise_()
        self.B_close.raise_()
        Result.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(Result)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 758, 26))
        self.menubar.setObjectName("menubar")
        Result.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(Result)
        self.statusbar.setObjectName("statusbar")
        Result.setStatusBar(self.statusbar)

        self.retranslateUi(Result)
        QtCore.QMetaObject.connectSlotsByName(Result)

    def retranslateUi(self, Result):
        _translate = QtCore.QCoreApplication.translate
        Result.setWindowTitle(_translate("Result", "MainWindow"))
        self.L_id.setText(_translate("Result", "Your id:"))
        self.LE_id.setToolTip(_translate("Result", "Type your id"))
        self.L_result.setText(_translate("Result", "Result:"))
        self.B_view.setToolTip(_translate("Result", "click to view"))
        self.B_view.setText(_translate("Result", "View"))
        self.B_close.setToolTip(_translate("Result", "exit"))
        self.B_close.setText(_translate("Result", "Close"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Result = QtWidgets.QMainWindow()
    ui = Ui_Result()
    ui.setupUi(Result)
    Result.show()
    sys.exit(app.exec_())

