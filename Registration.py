# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Registration.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook
from PyQt5.QtWidgets import QMessageBox


class Ui_Registration(object):
    
    def submit(self):
        try:
            active=self.wb['Sheet1']
            r=0
            row=True
            while row:
                r+=1
                for c in range(3):
                    if (active.cell(row=r+1,column=c+1).value) is None:
                        active.cell(row=r+1,column=c+1).value=self.LE_name.text()
                        active.cell(row=r+1,column=c+2).value=self.LE_ph.text()
                        active.cell(row=r+1,column=c+3).value=self.LE_email.text()
                        row=False
                        break
                        
            self.wb.create_sheet(self.LE_name.text()+'({})'.format(self.L_id.text()),len(self.wb.sheetnames))
            self.wb.save(r'Users\Accounts.xlsx')
            msg=QMessageBox()
            msg.setWindowTitle('Info')
            msg.setText('Created, note your account number:{}'.format(len(self.wb.sheetnames)-1))
            msg.setIcon(QMessageBox.Information)
            x=msg.exec_()
        except:
            msg=QMessageBox()
            msg.setWindowTitle('Info')
            msg.setText('Error!')
            msg.setIcon(QMessageBox.Critical)
            x=msg.exec_()
            
    
    def setupUi(self, Registration):
        
        self.wb=load_workbook(r'Users\Accounts.xlsx')
        
        Registration.setObjectName("Registration")
        Registration.setFixedSize(628, 549)
        self.centralwidget = QtWidgets.QWidget(Registration)
        self.centralwidget.setObjectName("centralwidget")
        self.L_your_id = QtWidgets.QLabel(self.centralwidget)
        self.L_your_id.setGeometry(QtCore.QRect(116, 160, 241, 61))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.L_your_id.setFont(font)
        self.L_your_id.setObjectName("L_your_id")
        self.L_name = QtWidgets.QLabel(self.centralwidget)
        self.L_name.setGeometry(QtCore.QRect(116, 220, 241, 61))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.L_name.setFont(font)
        self.L_name.setObjectName("L_name")
        self.L_ph = QtWidgets.QLabel(self.centralwidget)
        self.L_ph.setGeometry(QtCore.QRect(116, 290, 241, 61))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.L_ph.setFont(font)
        self.L_ph.setObjectName("L_ph")
        self.L_email = QtWidgets.QLabel(self.centralwidget)
        self.L_email.setGeometry(QtCore.QRect(116, 360, 241, 61))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.L_email.setFont(font)
        self.L_email.setObjectName("L_email")
        self.LE_name = QtWidgets.QLineEdit(self.centralwidget)
        self.LE_name.setGeometry(QtCore.QRect(350, 240, 221, 22))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.LE_name.setFont(font)
        self.LE_name.setObjectName("LE_name")
        self.LE_ph = QtWidgets.QLineEdit(self.centralwidget)
        self.LE_ph.setGeometry(QtCore.QRect(350, 310, 221, 22))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.LE_ph.setFont(font)
        self.LE_ph.setObjectName("LE_ph")
        self.LE_email = QtWidgets.QLineEdit(self.centralwidget)
        self.LE_email.setGeometry(QtCore.QRect(350, 380, 221, 22))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.LE_email.setFont(font)
        self.LE_email.setObjectName("LE_email")
        self.L_id = QtWidgets.QLabel(self.centralwidget)
        self.L_id.setGeometry(QtCore.QRect(390, 170, 141, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.L_id.setFont(font)
        self.L_id.setText("")
        self.L_id.setObjectName("L_id")
        
        self.L_id.setText(str(len(self.wb.sheetnames)))
        
        self.L_welcome = QtWidgets.QLabel(self.centralwidget)
        self.L_welcome.setGeometry(QtCore.QRect(20, 20, 591, 101))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(24)
        font.setBold(True)
        font.setWeight(75)
        self.L_welcome.setFont(font)
        self.L_welcome.setAlignment(QtCore.Qt.AlignCenter)
        self.L_welcome.setObjectName("L_welcome")
        self.bg_image = QtWidgets.QLabel(self.centralwidget)
        self.bg_image.setGeometry(QtCore.QRect(10, 10, 611, 501))
        self.bg_image.setText("")
        self.bg_image.setPixmap(QtGui.QPixmap("Assets\\Registration\\background.png"))
        self.bg_image.setScaledContents(True)
        self.bg_image.setObjectName("bg_image")
        self.B_exit = QtWidgets.QPushButton(self.centralwidget)
        self.B_exit.setGeometry(QtCore.QRect(430, 460, 71, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        self.B_exit.setFont(font)
        self.B_exit.setObjectName("B_exit")
        
        self.B_exit.clicked.connect(Registration.close)
        
        self.B_submit = QtWidgets.QPushButton(self.centralwidget)
        self.B_submit.setGeometry(QtCore.QRect(510, 460, 101, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        self.B_submit.setFont(font)
        self.B_submit.setObjectName("B_submit")
        
        self.B_submit.clicked.connect(self.submit)
        
        self.bg_image.raise_()
        self.L_your_id.raise_()
        self.L_name.raise_()
        self.L_ph.raise_()
        self.L_email.raise_()
        self.LE_name.raise_()
        self.LE_ph.raise_()
        self.LE_email.raise_()
        self.L_id.raise_()
        self.L_welcome.raise_()
        self.B_exit.raise_()
        self.B_submit.raise_()
        Registration.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(Registration)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 628, 26))
        self.menubar.setObjectName("menubar")
        Registration.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(Registration)
        self.statusbar.setObjectName("statusbar")
        Registration.setStatusBar(self.statusbar)

        self.retranslateUi(Registration)
        QtCore.QMetaObject.connectSlotsByName(Registration)

    def retranslateUi(self, Registration):
        _translate = QtCore.QCoreApplication.translate
        Registration.setWindowTitle(_translate("Registration", "Register"))
        self.L_your_id.setText(_translate("Registration", "Your ID"))
        self.L_name.setText(_translate("Registration", "Name"))
        self.L_ph.setText(_translate("Registration", "ph no.:"))
        self.L_email.setText(_translate("Registration", "Email id"))
        self.LE_name.setToolTip(_translate("Registration", "Type your name"))
        self.LE_ph.setToolTip(_translate("Registration", "Phone number"))
        self.LE_email.setToolTip(_translate("Registration", "Email address"))
        self.L_id.setToolTip(_translate("Registration", "Your unique ID number"))
        self.L_welcome.setText(_translate("Registration", "Welcome new user"))
        self.B_exit.setText(_translate("Registration", "exit"))
        self.B_submit.setText(_translate("Registration", "submit"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Registration = QtWidgets.QMainWindow()
    ui = Ui_Registration()
    ui.setupUi(Registration)
    Registration.show()
    sys.exit(app.exec_())

