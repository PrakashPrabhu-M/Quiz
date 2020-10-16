# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'testui.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

'''simple_peer peer_js'''

from PyQt5 import QtCore, QtGui, QtWidgets
import pandas as pd,random
from PyQt5.QtWidgets import QMessageBox
from openpyxl import load_workbook
import threading
import time
import datetime as dt

class Ui_Test(object):
    def counter(self):
        for i in range(60,-1,-1):
            time.sleep(1)
            print(f'\r{i}')
            self.L_timeRemaining.setText(f'\r{i}')
        self.sub()
        self.t1._stop()
    
    def sub(self):
        try:
            self.wb=load_workbook(r'Users\Accounts.xlsx')
            active=self.wb[self.wb.sheetnames[self.id]]
            active.cell(row=1,column=1).value='Test Category'
            active.cell(row=1,column=2).value='Score'   
            active.cell(row=1,column=3).value='Date'
            active.cell(row=1,column=4).value='Time'
            
            for i in self.mcq:
                for j in i:
                    if j.isChecked():
                        continue
                        #print(j.text())
            #print([i.toPlainText() for i in self.fill])
            #print(self.cato[self.rQ][0])
            m=-1
            f=-1
            #print(self.cato)
            
            for i in self.rQ:
                if self.cato[i]=='mcq':
                    an=None
                    m+=1
                    for j in self.mcq[m]:
                        if j.isChecked():
                            an=j.text()
                    #print('got an',an)
                    #print('answer',self.answers[i])
                    if an==self.answers[i]:
                        self.score+=1
                else:
                    f+=1
                    #print(self.answers[i],self.fill[f].toPlainText())
                    if self.answers[i]==self.fill[f].toPlainText():
                        self.score+=1
            
            #print(self.score)
            
            row=True
            r=0
            now=dt.datetime.now()
            while row:
                r+=1
                for c in range(2):
                    if active.cell(row=r+1,column=c+1).value is None:
                        active.cell(row=r+1,column=c+1).value=self.ques
                        active.cell(row=r+1,column=c+2).value=self.score
                        active.cell(row=r+1,column=c+3).value=f'{now.day}\\{now.month}\\{now.year}'
                        active.cell(row=r+1,column=c+4).value=f'{now.hour}:{now.minute}:{now.second}'
                        row=False
                        break
            self.wb.save(r'Users\Accounts.xlsx')
            msg=QMessageBox()
            msg.setWindowTitle('Info')
            msg.setText(f'You may go\nYour score: {self.score}')
            msg.setIcon(QMessageBox.Information)
            x=msg.exec_()
        except:
            msg=QMessageBox()
            msg.setWindowTitle('Info')
            msg.setText('Error!')
            msg.setIcon(QMessageBox.Critical)
            x=msg.exec_()
        
    def setupUi(self, Test,quesType='Category_1',rnum=1):  
        self.t1=threading.Thread(target=self.counter)
        self.t1.start()  
        print(f'[ACTIVE THREADS] {threading.activeCount()}')
        self.ques=quesType
        self.id=int(rnum)
        self.nQ=0
        self.score=0
        self.mcq=[]
        self.fill=[]
        self.data=pd.read_excel(r'Test contents\{}\Question_contents.xlsx'.format(quesType))
        a=[i for i in range(len(self.data))]
        random.shuffle(a)
        self.rQ=a[:10]
        self.questions=self.data.loc[self.rQ]["Questions"]
        self.answers=self.data.loc[self.rQ]["Answers"]
        self.cato=self.data.loc[self.rQ]["Type"]
        self.choice=self.data.loc[self.rQ][["Choice1",'Choice2','Choice3','Choice4']]
        
        Test.setObjectName("Test")
        Test.setFixedSize(897, 705)
        self.centralwidget = QtWidgets.QWidget(Test)
        self.centralwidget.setObjectName("centralwidget")
        self.scrollArea = QtWidgets.QScrollArea(self.centralwidget)
        self.scrollArea.setGeometry(QtCore.QRect(180, 70, 691, 531))
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, 668, 2505))
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.scrollAreaWidgetContents)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.frame = QtWidgets.QFrame(self.scrollAreaWidgetContents)
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.frame)
        self.verticalLayout.setObjectName("verticalLayout")
        self.L_q1 = QtWidgets.QLabel(self.frame)
        self.L_q1.setObjectName("L_q1")
        self.verticalLayout.addWidget(self.L_q1)    
        
        if self.cato[self.rQ[0]]!='mcq':
            self.tE_a1 = QtWidgets.QTextEdit(self.frame)        
            self.tE_a1.setObjectName("tE_a1")
            self.verticalLayout.addWidget(self.tE_a1)
            
            self.fill.append(self.tE_a1)
            
        else:
            self.rB_11 = QtWidgets.QRadioButton(self.frame)
            self.rB_11.setObjectName("rB_11")
            self.verticalLayout.addWidget(self.rB_11)
            self.rB_12 = QtWidgets.QRadioButton(self.frame)
            self.rB_12.setObjectName("rB_12")
            self.verticalLayout.addWidget(self.rB_12)
            self.rB_13 = QtWidgets.QRadioButton(self.frame)
            self.rB_13.setObjectName("rB_13")
            self.verticalLayout.addWidget(self.rB_13)
            self.rB_14 = QtWidgets.QRadioButton(self.frame)
            self.rB_14.setObjectName("rB_14")
            self.verticalLayout.addWidget(self.rB_14)
            self.mcq.append([self.rB_11,self.rB_12,self.rB_13,self.rB_14])
            
            
        self.frame_2 = QtWidgets.QFrame(self.frame)
        self.frame_2.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_2.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_2.setObjectName("frame_2")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.frame_2)
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.L_q2 = QtWidgets.QLabel(self.frame_2)
        self.L_q2.setObjectName("L_q2")
        self.verticalLayout_3.addWidget(self.L_q2)
        
        if self.cato[self.rQ[1]]!='mcq':
            self.tE_a2 = QtWidgets.QTextEdit(self.frame_2)
            self.tE_a2.setObjectName("tE_a2")
            self.verticalLayout_3.addWidget(self.tE_a2)
            
            self.fill.append(self.tE_a2)
            
            
        else:
            self.rB_21 = QtWidgets.QRadioButton(self.frame_2)
            self.rB_21.setObjectName("rB_21")
            self.verticalLayout_3.addWidget(self.rB_21)
            self.rB_22 = QtWidgets.QRadioButton(self.frame_2)
            self.rB_22.setObjectName("rB_22")
            self.verticalLayout_3.addWidget(self.rB_22)
            self.rB_23 = QtWidgets.QRadioButton(self.frame_2)
            self.rB_23.setObjectName("rB_23")
            self.verticalLayout_3.addWidget(self.rB_23)
            self.rB_24 = QtWidgets.QRadioButton(self.frame_2)
            self.rB_24.setObjectName("rB_24")
            self.verticalLayout_3.addWidget(self.rB_24)
            self.mcq.append([self.rB_21,self.rB_22,self.rB_23,self.rB_24])
            
        self.verticalLayout.addWidget(self.frame_2)
        self.verticalLayout_2.addWidget(self.frame)
        self.frame_3 = QtWidgets.QFrame(self.scrollAreaWidgetContents)
        self.frame_3.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_3.setObjectName("frame_3")
        self.verticalLayout_4 = QtWidgets.QVBoxLayout(self.frame_3)
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        self.L_q3 = QtWidgets.QLabel(self.frame_3)
        self.L_q3.setObjectName("L_q3")
        self.verticalLayout_4.addWidget(self.L_q3)
        
        if self.cato[self.rQ[2]]!='mcq':
            self.tE_a3 = QtWidgets.QTextEdit(self.frame_3)
            self.tE_a3.setObjectName("tE_a3")
            self.verticalLayout_4.addWidget(self.tE_a3)
            
            self.fill.append(self.tE_a3)
            
            
        else:
            self.rB_31 = QtWidgets.QRadioButton(self.frame_3)
            self.rB_31.setObjectName("rB_31")
            self.verticalLayout_4.addWidget(self.rB_31)
            self.rB_32 = QtWidgets.QRadioButton(self.frame_3)
            self.rB_32.setObjectName("rB_32")
            self.verticalLayout_4.addWidget(self.rB_32)
            self.rB_33 = QtWidgets.QRadioButton(self.frame_3)
            self.rB_33.setObjectName("rB_33")
            self.verticalLayout_4.addWidget(self.rB_33)
            self.rB_34 = QtWidgets.QRadioButton(self.frame_3)
            self.rB_34.setObjectName("rB_34")
            self.verticalLayout_4.addWidget(self.rB_34)
            self.mcq.append([self.rB_31,self.rB_32,self.rB_33,self.rB_34])
            
        self.verticalLayout_2.addWidget(self.frame_3)
        self.frame_4 = QtWidgets.QFrame(self.scrollAreaWidgetContents)
        self.frame_4.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_4.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_4.setObjectName("frame_4")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout(self.frame_4)
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.L_q4 = QtWidgets.QLabel(self.frame_4)
        self.L_q4.setObjectName("L_q4")
        self.verticalLayout_5.addWidget(self.L_q4)
        
        if self.cato[self.rQ[3]]!='mcq':
            self.tE_a4 = QtWidgets.QTextEdit(self.frame_4)
            self.tE_a4.setObjectName("tE_a4")
            self.verticalLayout_5.addWidget(self.tE_a4)
            
            self.fill.append(self.tE_a4)
            
            
        else:
            self.rB_41 = QtWidgets.QRadioButton(self.frame_4)
            self.rB_41.setObjectName("rB_41")
            self.verticalLayout_5.addWidget(self.rB_41)
            self.rB_42 = QtWidgets.QRadioButton(self.frame_4)
            self.rB_42.setObjectName("rB_42")
            self.verticalLayout_5.addWidget(self.rB_42)
            self.rB_43 = QtWidgets.QRadioButton(self.frame_4)
            self.rB_43.setObjectName("rB_43")
            self.verticalLayout_5.addWidget(self.rB_43)
            self.rB_44 = QtWidgets.QRadioButton(self.frame_4)
            self.rB_44.setObjectName("rB_44")
            self.verticalLayout_5.addWidget(self.rB_44)
            self.mcq.append([self.rB_41,self.rB_42,self.rB_43,self.rB_44])
            
        self.verticalLayout_2.addWidget(self.frame_4)
        self.frame_5 = QtWidgets.QFrame(self.scrollAreaWidgetContents)
        self.frame_5.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_5.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_5.setObjectName("frame_5")
        self.verticalLayout_6 = QtWidgets.QVBoxLayout(self.frame_5)
        self.verticalLayout_6.setObjectName("verticalLayout_6")
        self.L_q5 = QtWidgets.QLabel(self.frame_5)
        self.L_q5.setObjectName("L_q5")
        self.verticalLayout_6.addWidget(self.L_q5)
        
        if self.cato[self.rQ[4]]!='mcq':
            self.tE_a5 = QtWidgets.QTextEdit(self.frame_5)
            self.tE_a5.setObjectName("tE_a5")
            self.verticalLayout_6.addWidget(self.tE_a5)
            
            self.fill.append(self.tE_a5)
            
            
        else:
            self.rB_51 = QtWidgets.QRadioButton(self.frame_5)
            self.rB_51.setObjectName("rB_51")
            self.verticalLayout_6.addWidget(self.rB_51)
            self.rB_52 = QtWidgets.QRadioButton(self.frame_5)
            self.rB_52.setObjectName("rB_52")
            self.verticalLayout_6.addWidget(self.rB_52)
            self.rB_53 = QtWidgets.QRadioButton(self.frame_5)
            self.rB_53.setObjectName("rB_53")
            self.verticalLayout_6.addWidget(self.rB_53)
            self.rB_54 = QtWidgets.QRadioButton(self.frame_5)
            self.rB_54.setObjectName("rB_54")
            self.verticalLayout_6.addWidget(self.rB_54)
            self.mcq.append([self.rB_51,self.rB_52,self.rB_53,self.rB_54])
            
        self.verticalLayout_2.addWidget(self.frame_5)
        self.frame_6 = QtWidgets.QFrame(self.scrollAreaWidgetContents)
        self.frame_6.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_6.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_6.setObjectName("frame_6")
        self.verticalLayout_7 = QtWidgets.QVBoxLayout(self.frame_6)
        self.verticalLayout_7.setObjectName("verticalLayout_7")
        self.L_q6 = QtWidgets.QLabel(self.frame_6)
        self.L_q6.setObjectName("L_q6")
        self.verticalLayout_7.addWidget(self.L_q6)
        
        if self.cato[self.rQ[5]]!='mcq':
            self.tE_a6 = QtWidgets.QTextEdit(self.frame_6)
            self.tE_a6.setObjectName("tE_a6")
            self.verticalLayout_7.addWidget(self.tE_a6)
            
            self.fill.append(self.tE_a6)
            
            
        else:
            self.rB_61 = QtWidgets.QRadioButton(self.frame_6)
            self.rB_61.setObjectName("rB_61")
            self.verticalLayout_7.addWidget(self.rB_61)
            self.rB_62 = QtWidgets.QRadioButton(self.frame_6)
            self.rB_62.setObjectName("rB_62")
            self.verticalLayout_7.addWidget(self.rB_62)
            self.rB_63 = QtWidgets.QRadioButton(self.frame_6)
            self.rB_63.setObjectName("rB_63")
            self.verticalLayout_7.addWidget(self.rB_63)
            self.rB_64 = QtWidgets.QRadioButton(self.frame_6)
            self.rB_64.setObjectName("rB_64")
            self.verticalLayout_7.addWidget(self.rB_64)
            self.mcq.append([self.rB_61,self.rB_62,self.rB_63,self.rB_64])
            
        self.verticalLayout_2.addWidget(self.frame_6)
        self.frame_7 = QtWidgets.QFrame(self.scrollAreaWidgetContents)
        self.frame_7.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_7.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_7.setObjectName("frame_7")
        self.verticalLayout_8 = QtWidgets.QVBoxLayout(self.frame_7)
        self.verticalLayout_8.setObjectName("verticalLayout_8")
        self.L_q7 = QtWidgets.QLabel(self.frame_7)
        self.L_q7.setObjectName("L_q7")
        self.verticalLayout_8.addWidget(self.L_q7)
        
        if self.cato[self.rQ[6]]!='mcq':
            self.tE_a7 = QtWidgets.QTextEdit(self.frame_7)
            self.tE_a7.setObjectName("tE_a7")
            self.verticalLayout_8.addWidget(self.tE_a7)
            
            self.fill.append(self.tE_a7)
            
            
        else:
            self.rB_71 = QtWidgets.QRadioButton(self.frame_7)
            self.rB_71.setObjectName("rB_71")
            self.verticalLayout_8.addWidget(self.rB_71)
            self.rB_72 = QtWidgets.QRadioButton(self.frame_7)
            self.rB_72.setObjectName("rB_72")
            self.verticalLayout_8.addWidget(self.rB_72)
            self.rB_73 = QtWidgets.QRadioButton(self.frame_7)
            self.rB_73.setObjectName("rB_73")
            self.verticalLayout_8.addWidget(self.rB_73)
            self.rB_74 = QtWidgets.QRadioButton(self.frame_7)
            self.rB_74.setObjectName("rB_74")
            self.verticalLayout_8.addWidget(self.rB_74)
            self.mcq.append([self.rB_71,self.rB_72,self.rB_73,self.rB_74])
            
        self.verticalLayout_2.addWidget(self.frame_7)
        self.frame_8 = QtWidgets.QFrame(self.scrollAreaWidgetContents)
        self.frame_8.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_8.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_8.setObjectName("frame_8")
        self.verticalLayout_9 = QtWidgets.QVBoxLayout(self.frame_8)
        self.verticalLayout_9.setObjectName("verticalLayout_9")
        self.L_q8 = QtWidgets.QLabel(self.frame_8)
        self.L_q8.setObjectName("L_q8")
        self.verticalLayout_9.addWidget(self.L_q8)
        
        if self.cato[self.rQ[7]]!='mcq':
            self.tE_a8 = QtWidgets.QTextEdit(self.frame_8)
            self.tE_a8.setObjectName("tE_a8")
            self.verticalLayout_9.addWidget(self.tE_a8)
            
            self.fill.append(self.tE_a8)
            
            
        else:
            self.rB_81 = QtWidgets.QRadioButton(self.frame_8)
            self.rB_81.setObjectName("rB_81")
            self.verticalLayout_9.addWidget(self.rB_81)
            self.rB_82 = QtWidgets.QRadioButton(self.frame_8)
            self.rB_82.setObjectName("rB_82")
            self.verticalLayout_9.addWidget(self.rB_82)
            self.rB_83 = QtWidgets.QRadioButton(self.frame_8)
            self.rB_83.setObjectName("rB_83")
            self.verticalLayout_9.addWidget(self.rB_83)
            self.rB_84 = QtWidgets.QRadioButton(self.frame_8)
            self.rB_84.setObjectName("rB_84")
            self.verticalLayout_9.addWidget(self.rB_84)
            self.mcq.append([self.rB_81,self.rB_82,self.rB_83,self.rB_84])
            
        self.verticalLayout_2.addWidget(self.frame_8)
        self.frame_9 = QtWidgets.QFrame(self.scrollAreaWidgetContents)
        self.frame_9.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_9.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_9.setObjectName("frame_9")
        self.verticalLayout_10 = QtWidgets.QVBoxLayout(self.frame_9)
        self.verticalLayout_10.setObjectName("verticalLayout_10")
        self.L_q9 = QtWidgets.QLabel(self.frame_9)
        self.L_q9.setObjectName("L_q9")
        self.verticalLayout_10.addWidget(self.L_q9)
        
        if self.cato[self.rQ[8]]!='mcq':
            self.tE_a9 = QtWidgets.QTextEdit(self.frame_9)
            self.tE_a9.setObjectName("tE_a9")
            self.verticalLayout_10.addWidget(self.tE_a9)
            
            self.fill.append(self.tE_a9)
            
            
        else:
            self.rB_91 = QtWidgets.QRadioButton(self.frame_9)
            self.rB_91.setObjectName("rB_91")
            self.verticalLayout_10.addWidget(self.rB_91)
            self.rB_92 = QtWidgets.QRadioButton(self.frame_9)
            self.rB_92.setObjectName("rB_92")
            self.verticalLayout_10.addWidget(self.rB_92)
            self.rB_93 = QtWidgets.QRadioButton(self.frame_9)
            self.rB_93.setObjectName("rB_93")
            self.verticalLayout_10.addWidget(self.rB_93)
            self.rB_94 = QtWidgets.QRadioButton(self.frame_9)
            self.rB_94.setObjectName("rB_94")
            self.verticalLayout_10.addWidget(self.rB_94)
            self.mcq.append([self.rB_91,self.rB_92,self.rB_93,self.rB_94])
            
        self.verticalLayout_2.addWidget(self.frame_9)
        self.frame_10 = QtWidgets.QFrame(self.scrollAreaWidgetContents)
        self.frame_10.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_10.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_10.setObjectName("frame_10")
        self.verticalLayout_11 = QtWidgets.QVBoxLayout(self.frame_10)
        self.verticalLayout_11.setObjectName("verticalLayout_11")
        self.L_q10 = QtWidgets.QLabel(self.frame_10)
        self.L_q10.setObjectName("L_q10")
        self.verticalLayout_11.addWidget(self.L_q10)
        
        if self.cato[self.rQ[9]]!='mcq':
            self.tE_a10 = QtWidgets.QTextEdit(self.frame_10)
            self.tE_a10.setObjectName("tE_a10")
            self.verticalLayout_11.addWidget(self.tE_a10)
            
            self.fill.append(self.tE_a10)
            
            
        else:
            self.rB_101 = QtWidgets.QRadioButton(self.frame_10)
            self.rB_101.setObjectName("rB_101")
            self.verticalLayout_11.addWidget(self.rB_101)
            self.rB_102 = QtWidgets.QRadioButton(self.frame_10)
            self.rB_102.setObjectName("rB_102")
            self.verticalLayout_11.addWidget(self.rB_102)
            self.rB_103 = QtWidgets.QRadioButton(self.frame_10)
            self.rB_103.setObjectName("rB_103")
            self.verticalLayout_11.addWidget(self.rB_103)
            self.rB_104 = QtWidgets.QRadioButton(self.frame_10)
            self.rB_104.setObjectName("rB_104")
            self.verticalLayout_11.addWidget(self.rB_104)
            self.mcq.append([self.rB_101,self.rB_102,self.rB_103,self.rB_104])
            
        self.verticalLayout_2.addWidget(self.frame_10)
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.B_submit = QtWidgets.QPushButton(self.centralwidget)
        self.B_submit.setGeometry(QtCore.QRect(780, 610, 93, 28))
        self.B_submit.setObjectName("B_submit")
        
        self.B_submit.clicked.connect(self.sub)
        
        self.B_cancel = QtWidgets.QPushButton(self.centralwidget)
        self.B_cancel.setGeometry(QtCore.QRect(680, 610, 93, 28))
        self.B_cancel.setObjectName("B_cancel")
        
        self.B_cancel.clicked.connect(Test.close)
        
        self.L_quiz = QtWidgets.QLabel(self.centralwidget)
        self.L_quiz.setGeometry(QtCore.QRect(184, 10, 681, 41))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(21)
        font.setBold(True)
        font.setItalic(True)
        font.setWeight(75)
        self.L_quiz.setFont(font)
        self.L_quiz.setAlignment(QtCore.Qt.AlignCenter)
        self.L_quiz.setObjectName("L_quiz")
        self.bg_image = QtWidgets.QLabel(self.centralwidget)
        self.bg_image.setGeometry(QtCore.QRect(0, -5, 891, 671))
        self.bg_image.setText("")
        self.bg_image.setPixmap(QtGui.QPixmap(r"Assets\Take test\background.png"))
        self.bg_image.setScaledContents(True)
        self.bg_image.setObjectName("bg_image")
        self.L_time = QtWidgets.QLabel(self.centralwidget)
        self.L_time.setGeometry(QtCore.QRect(10, 20, 171, 51))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.L_time.setFont(font)
        self.L_time.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignVCenter)
        self.L_time.setObjectName("L_time")
        
        self.L_timeRemaining = QtWidgets.QLabel(self.centralwidget)
        self.L_timeRemaining.setGeometry(QtCore.QRect(184, 19, 111, 51))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.L_timeRemaining.setFont(font)
        self.L_timeRemaining.setText("")
        self.L_timeRemaining.setAlignment(QtCore.Qt.AlignCenter)
        self.L_timeRemaining.setObjectName("L_timeRemaining")
        self.bg_image.raise_()
        self.scrollArea.raise_()
        self.B_submit.raise_()
        self.B_cancel.raise_()
        self.L_quiz.raise_()
        self.L_time.raise_()
        self.L_timeRemaining.raise_()
        Test.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(Test)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 897, 26))
        self.menubar.setObjectName("menubar")
        Test.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(Test)
        self.statusbar.setObjectName("statusbar")
        Test.setStatusBar(self.statusbar)

        self.retranslateUi(Test)
        QtCore.QMetaObject.connectSlotsByName(Test)

    def retranslateUi(self, Test):
        _translate = QtCore.QCoreApplication.translate
        Test.setWindowTitle(_translate("Test", "Exam"))
        
        self.nQ+=1
        self.L_q1.setText(_translate("Test", str(self.nQ)+') '+str(self.questions[self.rQ[0]])))
        if self.cato[self.rQ[0]]=='mcq':
            self.rB_11.setText(_translate("Test", self.choice['Choice1'][self.rQ[0]]))
            self.rB_12.setText(_translate("Test", self.choice['Choice2'][self.rQ[0]]))
            self.rB_13.setText(_translate("Test", self.choice['Choice3'][self.rQ[0]]))
            self.rB_14.setText(_translate("Test", self.choice['Choice4'][self.rQ[0]]))
        
        self.nQ+=1
        self.L_q2.setText(_translate("Test", str(self.nQ)+') '+str(self.questions[self.rQ[1]])))
        if self.cato[self.rQ[1]]=='mcq':
            self.rB_21.setText(_translate("Test", self.choice['Choice1'][self.rQ[1]]))
            self.rB_22.setText(_translate("Test", self.choice['Choice2'][self.rQ[1]]))
            self.rB_23.setText(_translate("Test", self.choice['Choice3'][self.rQ[1]]))
            self.rB_24.setText(_translate("Test", self.choice['Choice4'][self.rQ[1]]))
        
        self.nQ+=1
        self.L_q3.setText(_translate("Test", str(self.nQ)+') '+str(self.questions[self.rQ[2]])))
        if self.cato[self.rQ[2]]=='mcq':
            self.rB_31.setText(_translate("Test", self.choice['Choice1'][self.rQ[2]]))
            self.rB_32.setText(_translate("Test", self.choice['Choice2'][self.rQ[2]]))
            self.rB_33.setText(_translate("Test", self.choice['Choice3'][self.rQ[2]]))
            self.rB_34.setText(_translate("Test", self.choice['Choice4'][self.rQ[2]]))
        
        self.nQ+=1 
        self.L_q4.setText(_translate("Test", str(self.nQ)+') '+str(self.questions[self.rQ[3]])))
        if self.cato[self.rQ[3]]=='mcq':
            self.rB_41.setText(_translate("Test", self.choice['Choice1'][self.rQ[3]]))
            self.rB_42.setText(_translate("Test", self.choice['Choice2'][self.rQ[3]]))
            self.rB_43.setText(_translate("Test", self.choice['Choice3'][self.rQ[3]]))
            self.rB_44.setText(_translate("Test", self.choice['Choice4'][self.rQ[3]]))
        
        self.nQ+=1
        self.L_q5.setText(_translate("Test", str(self.nQ)+') '+str(self.questions[self.rQ[4]])))
        if self.cato[self.rQ[4]]=='mcq':
            self.rB_51.setText(_translate("Test", self.choice['Choice1'][self.rQ[4]]))
            self.rB_52.setText(_translate("Test", self.choice['Choice2'][self.rQ[4]]))
            self.rB_53.setText(_translate("Test", self.choice['Choice3'][self.rQ[4]]))
            self.rB_54.setText(_translate("Test", self.choice['Choice4'][self.rQ[4]]))
        
        self.nQ+=1
        self.L_q6.setText(_translate("Test", str(self.nQ)+') '+str(self.questions[self.rQ[5]])))
        if self.cato[self.rQ[5]]=='mcq':
            self.rB_61.setText(_translate("Test", self.choice['Choice1'][self.rQ[5]]))
            self.rB_62.setText(_translate("Test", self.choice['Choice2'][self.rQ[5]]))
            self.rB_63.setText(_translate("Test", self.choice['Choice3'][self.rQ[5]]))
            self.rB_64.setText(_translate("Test", self.choice['Choice4'][self.rQ[5]]))
            
        self.nQ+=1
        self.L_q7.setText(_translate("Test", str(self.nQ)+') '+str(self.questions[self.rQ[6]])))
        if self.cato[self.rQ[6]]=='mcq':
            self.rB_71.setText(_translate("Test", self.choice['Choice1'][self.rQ[6]]))
            self.rB_72.setText(_translate("Test", self.choice['Choice2'][self.rQ[6]]))
            self.rB_73.setText(_translate("Test", self.choice['Choice3'][self.rQ[6]]))
            self.rB_74.setText(_translate("Test", self.choice['Choice4'][self.rQ[6]]))
            
        self.nQ+=1   
        self.L_q8.setText(_translate("Test", str(self.nQ)+') '+str(self.questions[self.rQ[7]])))
        if self.cato[self.rQ[7]]=='mcq':
            self.rB_81.setText(_translate("Test", self.choice['Choice1'][self.rQ[7]]))
            self.rB_82.setText(_translate("Test", self.choice['Choice2'][self.rQ[7]]))
            self.rB_83.setText(_translate("Test", self.choice['Choice3'][self.rQ[7]]))
            self.rB_84.setText(_translate("Test", self.choice['Choice4'][self.rQ[7]]))
            
        self.nQ+=1    
        self.L_q9.setText(_translate("Test", str(self.nQ)+') '+str(self.questions[self.rQ[8]])))
        if self.cato[self.rQ[8]]=='mcq':
            self.rB_91.setText(_translate("Test", self.choice['Choice1'][self.rQ[8]]))
            self.rB_92.setText(_translate("Test", self.choice['Choice2'][self.rQ[8]]))
            self.rB_93.setText(_translate("Test", self.choice['Choice3'][self.rQ[8]]))
            self.rB_94.setText(_translate("Test", self.choice['Choice4'][self.rQ[8]]))
            
        self.nQ+=1    
        self.L_q10.setText(_translate("Test", str(self.nQ)+') '+str(self.questions[self.rQ[9]])))
        if self.cato[self.rQ[9]]=='mcq':
            self.rB_101.setText(_translate("Test", self.choice['Choice1'][self.rQ[9]]))
            self.rB_102.setText(_translate("Test", self.choice['Choice2'][self.rQ[9]]))
            self.rB_103.setText(_translate("Test", self.choice['Choice3'][self.rQ[9]]))
            self.rB_104.setText(_translate("Test", self.choice['Choice4'][self.rQ[9]]))
            
            
        self.B_submit.setText(_translate("Test", "Submit"))
        self.B_cancel.setText(_translate("Test", "Cancel"))
        self.L_quiz.setText(_translate("Test", "Quiz"))
        self.L_time.setText(_translate("Test", "Time Remaining:"))
        
    def chosen(self,c):
        for i in c:
            if i.isChecked():
                return i.text()


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Test = QtWidgets.QMainWindow()
    ui = Ui_Test()
    ui.setupUi(Test)
    Test.show()
    sys.exit(app.exec_())

