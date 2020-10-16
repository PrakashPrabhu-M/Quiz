# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Authentication.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
from test2 import Ui_Test
import os
from openpyxl import load_workbook

class Ui_testAuthentication(object):
    
    def nxt(self):
        try:
            wb=load_workbook(r'Users\Accounts.xlsx')
            try:
                n=int(self.LE_id.text())
            except ValueError:
                n=9999999
            if n<=len(wb.sheetnames):
                self.window=QtWidgets.QMainWindow()
                self.ui=Ui_Test()
                quesType=self.comboBox.currentText()
                self.ui.setupUi(self.window,quesType,self.LE_id.text())
                #testAuthentication.hide()
                self.window.show()
            else:
                msg=QMessageBox()
                msg.setWindowTitle('Info')
                msg.setText('Invalid credentials')
                msg.setIcon(QMessageBox.Information)
                x=msg.exec_()
        except:
            msg=QMessageBox()
            msg.setWindowTitle('Info')
            msg.setText('Error!')
            msg.setIcon(QMessageBox.Critical)
            x=msg.exec_()
            
            
    
    def setupUi(self, testAuthentication):
        testAuthentication.setObjectName("testAuthentication")
        testAuthentication.setFixedSize(664, 214)
        self.centralwidget = QtWidgets.QWidget(testAuthentication)
        self.centralwidget.setObjectName("centralwidget")
        self.bg_image = QtWidgets.QLabel(self.centralwidget)
        self.bg_image.setGeometry(QtCore.QRect(0, 5, 661, 161))
        self.bg_image.setText("")
        self.bg_image.setPixmap(QtGui.QPixmap("Assets\\Take test\\background.png"))
        self.bg_image.setScaledContents(False)
        self.bg_image.setObjectName("bg_image")
        self.L_id = QtWidgets.QLabel(self.centralwidget)
        self.L_id.setGeometry(QtCore.QRect(150, 10, 161, 61))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.L_id.setFont(font)
        self.L_id.setObjectName("L_id")
        self.LE_id = QtWidgets.QLineEdit(self.centralwidget)
        self.LE_id.setGeometry(QtCore.QRect(310, 21, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.LE_id.setFont(font)
        self.LE_id.setObjectName("LE_id")
        self.B_cancel = QtWidgets.QPushButton(self.centralwidget)
        self.B_cancel.setGeometry(QtCore.QRect(320, 120, 93, 28))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(10)
        self.B_cancel.setFont(font)
        self.B_cancel.setObjectName("B_cancel")
        
        self.B_cancel.clicked.connect(testAuthentication.close)
        
        self.B_next = QtWidgets.QPushButton(self.centralwidget)
        self.B_next.setGeometry(QtCore.QRect(430, 120, 93, 28))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(10)
        self.B_next.setFont(font)
        self.B_next.setObjectName("B_next")
        
        self.B_next.clicked.connect(self.nxt)
        
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(310, 60, 221, 31))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.comboBox.setFont(font)
        self.comboBox.setObjectName("comboBox")
        
        for i in os.scandir('Test contents'):
            if i.is_dir():
                self.comboBox.addItem("")
        #self.comboBox.addItem("")
        #self.comboBox.addItem("")
        
        
        self.L_category = QtWidgets.QLabel(self.centralwidget)
        self.L_category.setGeometry(QtCore.QRect(150, 40, 161, 61))
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(14)
        self.L_category.setFont(font)
        self.L_category.setObjectName("L_category")
        testAuthentication.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(testAuthentication)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 664, 26))
        self.menubar.setObjectName("menubar")
        testAuthentication.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(testAuthentication)
        self.statusbar.setObjectName("statusbar")
        testAuthentication.setStatusBar(self.statusbar)
        
        self.retranslateUi(testAuthentication)
        QtCore.QMetaObject.connectSlotsByName(testAuthentication)

    def retranslateUi(self, testAuthentication):
        _translate = QtCore.QCoreApplication.translate
        testAuthentication.setWindowTitle(_translate("testAuthentication", "Authentication"))
        self.L_id.setText(_translate("testAuthentication", "Your id nmber:"))
        self.LE_id.setToolTip(_translate("testAuthentication", "Your id number"))
        self.B_cancel.setText(_translate("testAuthentication", "Cancel"))
        self.B_next.setText(_translate("testAuthentication", "Next >>"))
        self.comboBox.setToolTip(_translate("testAuthentication", "Choose the type of questions"))
        
        for n,i in enumerate(os.scandir('Test contents')):
            if i.is_dir():
                self.comboBox.setItemText(n, _translate("testAuthentication", i.name))
        # self.comboBox.setItemText(1, _translate("testAuthentication", "Type2"))
        # self.comboBox.setItemText(2, _translate("testAuthentication", "Type3"))
        
        self.L_category.setText(_translate("testAuthentication", "Category"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    testAuthentication = QtWidgets.QMainWindow()
    ui = Ui_testAuthentication()
    ui.setupUi(testAuthentication)
    testAuthentication.show()
    sys.exit(app.exec_())

